#!/usr/bin/env python
#-*- coding: utf-8 -*-

"""
  User localization via wifi attempts rssi
"""

import csv, os, re, sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.cell import get_column_letter

apinfo_wb = './data/apinfo.xlsx'
logpath = './data/tmplog/'

numofdays = 0
stat_days = None

stations = []

user_trace = {}

locinfo = {}

aplist = {}

daily_stat_conter = 0
total_counted_recs = 0
log_lines = 0

reload(sys)
sys.setdefaultencoding('utf8')

def init_daily_stat():
    stat = {}
    for day in stat_days:
        stat[day] = 0
    return stat

def get_ap_info(wbpath):
    global stations, user_trace
    wb = load_workbook(filename=wbpath)
    for sheetname in wb.sheetnames:
        station = {}
        station['name'] = sheetname
        user_trace[sheetname] = []
        sheet = wb.get_sheet_by_name(sheetname)
        for i in xrange(sheet.min_row, sheet.max_row+1):
            apname = macaddr = None
            apvalue = sheet.cell("A%d"%i).value
            macvalue = sheet.cell("C%d"%i).value
            if apvalue and macvalue:
                if apvalue.startswith("AP"):
                    if len(apvalue) < 4: apname = apvalue[:-1] + '0' + apvalue[-1]
                    else: apname = apvalue
                if '.' not in macvalue: macaddr = macvalue
                else: macaddr = sheet.cell("D%d"%i).value
            if apname and macaddr: 
                station[format_mac(macaddr)] = apname
        stations.append(station)
    print stations

def format_mac(ma):
    if '-' in ma:
        ma = ma.replace('-', '')
    elif ':' in ma:
        ma = ma.replace(':', '')
    return ':'.join([ma[i:i+2] for i in range(0, len(ma), 2)])

def get_aplist():
    global aplist
    for s in stations:
        for k in s.keys():
            if k.startswith('F') or k.startswith('8'):
                aplist[k] = s['name']+'-'+s[k]

class Device(object):
    def __init__(self, mac):
        self.mac = mac
        self.trace = []

    def add_point(self, connected_ap, loc, ts):
        self.trace.append(
            ( connected_ap, loc, int(ts), 
                datetime.strftime(datetime.fromtimestamp(int(ts)), '%Y-%m-%d %H:%M:%S'),
            )
        )

def group_daily_logs(path):
    """mkdir to group daily logs, 
       does't make sense to parse logs from different days 
       as no subway service during the night
    """
    logfiles = [log for log in os.listdir(path) if log.endswith('.log')]
    if logfiles:
        folders = set([log.split('.')[0][:-2] for log in logfiles])
        for fldr in folders:
            if not os.path.isdir(path+fldr):
                os.mkdir(path+fldr)
            for log in logfiles:
                if log.startswith(fldr):
                    os.rename(path+log, path+fldr+'/'+log)
    else: folders = [fld for fld in os.listdir(path) if not fld.startswith('.')]
    global numofdays
    numofdays = len(folders)
    print numofdays
    return folders

def parse(datafile):
    global log_lines
    fd = open(datafile, 'rb')
    for line in fd:
        log_lines += 1
        parse_line(line)
    fd.close()

def parse_line(ln):
    global locinfo
    split = [f.strip().upper() for f in ln.split(',')]
    if len(split) == 4 and not re.findall('[a-z,A-Z,:,-]', split[0]):
        ts, apmac, devmac, loc = tuple(split)
        rssi = 0
    elif len(split) == 5 and not re.findall('[a-z,A-Z,:,-]', split[0]):
        ts, apmac, devmac, rssi, loc = tuple(split)
    else:
        print ln
        return # invalid line, nth to do
    if locinfo.has_key(devmac):
        locinfo[devmac].add_point(aplist[apmac.upper()], loc, ts)
    else:
        locinfo[devmac] = Device(devmac)
        locinfo[devmac].add_point(aplist[apmac.upper()], loc, ts)

def reshape_trace(trace_pnts):
    traces = []
    res = []
    for p in trace_pnts:
        if res != []:
            lastp = res[-1]
            if (p[0][:-5] != lastp[0][:-5]) or (p[-2]-lastp[-2] > 300):
                #print p[0][:-5], '->', lastp[0][:-5]
                traces.append(res)
                res = []
        res.append(p)
    if res: # last trip
        traces.append(res)
    return traces

def raw_data(d):
    global locinfo
    locinfo = {}  #init for daily results
    logfiles = os.listdir(d)
    for log in [l for l in logfiles if l.endswith('.log')]:
        parse(d+'/'+log)
    num_of_points = 0
    for k, v in locinfo.items():
        num_of_points += len(v.trace)
    print "passed points: %d"%num_of_points

def get_route(user):
    mac = user.mac
    start = user.trace[0][0]
    end = user.trace[-1][0]
    enter_time = user.trace[0][-1]
    leave_time = user.trace[-1][-1]
    ontime, offtime = get_on_off(user.trace)
    if mac and start and end and enter_time and leave_time and ontime and offtime:
        return (mac.encode('utf8'), start.encode('utf8'), enter_time.encode('utf8'), ontime.encode('utf8'), 
            end.encode('utf8'), offtime.encode('utf8'), leave_time.encode('utf8'), form_route(user.trace).encode('utf8'))
    else:
        return None

def get_on_off(trace):
    for i in xrange(1, len(trace)):
        if trace[i][0][:-4] != trace[i-1][0][:-4]:
            if trace[i][-2] - trace[i-1][-2] > 120:
                return trace[i-1][-1], trace[i][-1]
    return 'na', 'na'

def is_fullpath(trace):
    is_start = is_end = False
    for i in xrange(1, len(trace)):
        if trace[i][0][:-4] != trace[i-1][0][:-4]:
            if trace[i-1][-2] - trace[0][-2] > 180:
                is_start = True
            break
    for i in xrange(1, len(trace)):
        if trace[0-i][0][:-4] != trace[0-i-1][0][:-4]:
            if trace[-1][-2] - trace[0-i][-2] > 180:
                is_end = True

    return (is_start, is_end)

def form_route(trace):
    return '->'.join([r[-1]+'@'+r[0] for r in trace])


def main():
    get_ap_info(apinfo_wb)
    get_aplist()
    #for mac, ap in aplist.items():
    #    print mac, ap
    raw_data()
    for k, v in locinfo.items():
        traces = reshape_trace(v.trace)
        for trace in traces:
            if trace[0][0][:-4] == trace[-1][0][:-4] and trace[-1][-2] - trace[0][-2] > 9000: 
                continue # invalid data if lasted for longer than 2.5 hours
            if trace[0][0][:-4] != trace[-1][0][:-4]: 
                isstart, isend = is_fullpath(trace)
                if isstart and isend:
                    route = get_route(v)
                    if route: 
                        #print route
                        results.append(route)
                elif isstart:
                    print "started from..."
                elif isend:
                    print "ended at..."
                else:
                    print "passed by at both..."
            else: # just passed by?
                if trace[-1][-2] - trace[0][-2] > 120 and re.findall('AP0[1-3]', trace[0][0]):
                    print "started journey @%s"%trace[0][0]
                elif trace[-1][-2] - trace[0][-2] > 120 and re.findall('AP0[1-3]', trace[-1][0]):
                    print "ended journey @%s"%trace[-1][0]
                else: print "passed by @ %s"%trace[0][0]
                #print k, ': ', form_route(v.trace) + '\n\n'
    store_results()

results = [] # (id(mac), start, enter_time, on_time, end, off_time, leave_time)
enteries = {}
exits = {}
passby = {}
fullpath = {}

def store_results():
    with open('detailed_results.csv', 'wb') as csvfile:
        writer = csv.writer(csvfile, dialect='excel')
        writer.writerow(('设备ID', '始发站', '进站时间', '上车时间', '终点站', '下车时间', '出站时间', '全称路线'))
        writer.writerows(results)

def create_stat_xls():
    wb = Workbook()
    ws_in_out = wb.create_sheet()
    ws_in_out.title = u"进出站统计"
    ws_fp = wb.create_sheet()
    ws_fp.title = u"起始终点统计"
    ws_pb = wb.create_sheet()
    ws_pb.title = u"途经站点统计"
    wb.save("stat_results.xlsx")


pb_results = {}
io_results = {}
fp_results = {}
passboth = {}

def open_result_wb():
    return load_workbook(filename='results.xlsx')

def update_passboth(trip, date):
    global passboth
    path = '%s-%s'%trip
    if passboth.has_key(path):
        passboth[path][date] += 1
    else:
        passboth[path] = init_daily_stat()
        passboth[path][date] += 1

def put_passboth(offset=0):
    for k, v in passboth.items():
        print k, ':', v
    wb = Workbook()
    ws = wb.active
    keys = passboth.keys()
    for i in range(len(keys)):
        ws['%s%d'%(get_column_letter(1), i+2)] = keys[i]
        days = passboth[keys[i]].keys()
        days.sort()
        for j in range(len(days)):
            ws['%s%d'%(get_column_letter(j+2), 1)] = days[j]
            ws['%s%d'%(get_column_letter(j+2), i+2)] = passboth[keys[i]][days[j]]
    wb.save('results.xlsx')
    return offset+len(keys)+1

def update_pb(station, date):
    global pb_results
    if pb_results.has_key(station):
        pb_results[station][date] += 1
    else:
        pb_results[station] = init_daily_stat()
        pb_results[station][date] += 1

def put_pb(offset=0):
    for k, v in pb_results.items():
        print k, ':', v
    wb = Workbook()
    ws = wb.active
    keys = pb_results.keys()
    for i in range(len(keys)):
        ws['%s%d'%(get_column_letter(1), i+2)] = keys[i]
        days = pb_results[keys[i]].keys()
        days.sort()
        for j in range(len(days)):
            ws['%s%d'%(get_column_letter(j+2), 1)] = days[j]
            ws['%s%d'%(get_column_letter(j+2), i+2)] = pb_results[keys[i]][days[j]]
    wb.save('pb_results.xlsx')
    return offset+len(keys)+1

def update_fp(trip, date):
    global fp_results
    path = '%s-%s'%trip
    if fp_results.has_key(path):
        fp_results[path][date] += 1
    else:
        fp_results[path] = init_daily_stat()
        fp_results[path][date] += 1

def put_fp(offset=0):
    for k, v in fp_results.items():
        print k, ':', v
    wb = Workbook()
    ws = wb.active
    keys = fp_results.keys()
    for i in range(len(keys)):
        ws['%s%d'%(get_column_letter(1), i+2)] = keys[i]
        days = fp_results[keys[i]].keys()
        days.sort()
        for j in range(len(days)):
            ws['%s%d'%(get_column_letter(j+2), 1)] = days[j]
            ws['%s%d'%(get_column_letter(j+2), i+2)] = fp_results[keys[i]][days[j]]
    wb.save('fp_results.xlsx')
    return offset+len(keys)+1

def update_io(station, on, date):
    global io_results
    if io_results.has_key(station):
        if on: io_results[station][date] += 1
        else:  io_results[station][date] += (1<<32)
    else:
        io_results[station] = init_daily_stat()
        if on: io_results[station][date] += 1
        else:  io_results[station][date] += (1<<32)

def put_io(offset=0):
    for k, v in io_results.items():
        print k, ':', v.values()[0]&0xffffffff, ' ', v.values()[0]>>32
    wb = Workbook()
    ws = wb.active
    keys = io_results.keys()
    for i in range(len(keys)):
        ws['%s%d'%(get_column_letter(1), i+2)] = keys[i]
        days = io_results[keys[i]].keys()
        days.sort()
        for j in range(len(days)):
            ws['%s%d'%(get_column_letter(2*j+2), 1)] = days[j]
            ws.merge_cells('%s%d:%s%d'%(get_column_letter(2*j+2), 1, get_column_letter(2*j+3), 1))
            ws['%s%d'%(get_column_letter(2*j+2), i+2)] = io_results[keys[i]][days[j]] & 0xffffffff
            ws['%s%d'%(get_column_letter(2*j+3), i+2)] = io_results[keys[i]][days[j]] >> 32
    wb.save('io_results.xlsx')
    return offset+len(keys)+1

#main()
if __name__ != '__main__':
    #group_daily_logs('./data1022/tmplog/')
    create_stat_xls()
else:
    try:
        get_ap_info(apinfo_wb)
        get_aplist()
    except:
        print 'failed to get AP...'

    global stat_days
    stat_days = group_daily_logs(logpath)
    for fldr in stat_days:
        print fldr
        raw_data(logpath+fldr)
        for k, v in locinfo.items():
            traces = reshape_trace(v.trace)
            for trace in traces:
                if trace[0][0][:-4] != trace[-1][0][:-4]:
                    print "never enter here"
                    isstart, isend = is_fullpath(trace)
                    if isstart and isend:
                        route = get_route(v)
                        if route: 
                            results.append(route)
                            update_fp((trace[0][0][:-5], trace[-1][0][:-5]), fldr)
                    elif isstart:
                        #print "started from %s"%trace[0][0][:-4]
                        update_io(trace[0][0][:-5], True, fldr)
                    elif isend:
                        #print "ended at %s"%trace[-1][0][:-4]
                        update_io(trace[-1][0][:-5], False, fldr)
                    else:
                        update_passboth((trace[0][0][:-5], trace[-1][0][:-5]), fldr)
                else: # just passed by?
                    global daily_stat_conter, total_counted_recs
                    daily_stat_conter += 1
                    total_counted_recs += len(trace)
                    #if trace[-1][-2] - trace[0][-2] > 9000: 
                    #    continue # invalid data if lasted for longer than 2.5 hours in same station
                    #if trace[-1][-2] - trace[0][-2] > 120 and u'站厅' in trace[0][1]:
                    #if trace[-1][-2] - trace[0][-2] > 120 and re.findall('AP0[1-3]', trace[0][0]):
                        #print "started journey @%s"%trace[0][0]
                    #    print '->'.join([p[1].strip('\n') for p in trace])
                    #    update_io(trace[0][0][:-5], True, fldr)
                    #elif trace[-1][-2] - trace[0][-2] > 120 and re.findall(u'站台', trace[-1][1]):
                        #print "ended journey @%s"%trace[-1][0]
                        #print '-'.join([p[1] for p in trace])
                    #    update_io(trace[-1][0][:-5], False, fldr)
                    #print '%s::'%k, '->'.join([p[-1]+'@'+p[1].strip('\n') for p in trace]), '\n'
                    #print [p[0][:-5] for p in trace], '\n'
                    if trace[-1][-2] - trace[0][-2] > 60:
                        if [t for t in trace if u'站台' in t[1].decode('utf8')]:
                            update_io(trace[-1][0][:-5], True, fldr)
                        else:
                            update_io(trace[-1][0][:-5], False, fldr) # not enter station
                    else:
                        if [t for t in trace if u'站台' in t[1].decode('utf8')]:
                            update_pb(trace[0][0][:-5], fldr)
                        else:
                            update_io(trace[0][0][:-5], False, fldr) # not enter station

    print 'passby:'; put_pb()
    #print 'fullpath:'; put_fp()
    print 'in out:'; put_io()
    #print 'both passby:'; put_passboth()
    print 'total log lines: %d'%log_lines
    print 'total recs: %d'%daily_stat_conter
    print 'total_counted_recs: %d'%total_counted_recs




